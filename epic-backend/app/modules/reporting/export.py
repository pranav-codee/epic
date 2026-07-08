"""
Excel and PDF export builders for the admin dashboard (REQ: dashboard data export).

Both builders reuse `service.overview()` so the exported figures always match what the
dashboard UI is showing. Kept dependency-light and synchronous — dashboard-sized datasets
(thousands of tickets, not millions) build in well under a second.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session, joinedload

from . import service
from ..tickets.models import Ticket

BRAND_HEX = "0067B8"
BORDER_HEX = "E1E4E8"
MUTED_HEX = "57606A"

HEADER_FILL = PatternFill(start_color=BRAND_HEX, end_color=BRAND_HEX, fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
MUTED_FONT = Font(color=MUTED_HEX, italic=True)

MAX_TICKET_ROWS = 5000  # sane ceiling for the raw ticket export sheet


# --------------------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------------------

def _autosize(ws):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length, 10), 50)


def _header_row(ws, row_idx: int):
    for cell in ws[row_idx]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _breakdown_sheet(wb, name: str, mapping: dict):
    ws = wb.create_sheet(name)
    ws.append(["Key", "Count"])
    _header_row(ws, 1)
    for k, v in mapping.items():
        ws.append([k, v])
    _autosize(ws)


def build_excel(db: Session) -> io.BytesIO:
    data = service.overview(db)
    sla = data["sla"]

    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws.append(["EPIC Helpdesk — Dashboard Export"])
    ws["A1"].font = TITLE_FONT
    ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
    ws["A2"].font = MUTED_FONT
    ws.append([])
    ws.append(["Metric", "Value"])
    _header_row(ws, 4)
    for k, v in [
        ("Total tickets", data["total_tickets"]),
        ("Open tickets", data["open_tickets"]),
        ("SLA compliance %", sla["compliance_pct"] if sla["compliance_pct"] is not None else "—"),
        ("SLA breached (resolved late)", sla["breached_count"]),
        ("Open tickets breaching SLA", sla["open_breached_count"]),
        ("Open tickets at risk", sla["open_at_risk_count"]),
        ("Avg resolution time (hrs)", sla["avg_resolution_hours_overall"] or "—"),
    ]:
        ws.append([k, v])
    _autosize(ws)

    _breakdown_sheet(wb, "By Status", data["by_status"])
    _breakdown_sheet(wb, "By Priority", data["by_priority"])
    _breakdown_sheet(wb, "By Category", data["by_category"])
    _breakdown_sheet(wb, "By Type", data["by_ticket_type"])

    # --- SLA sheet ---
    sla_ws = wb.create_sheet("SLA")
    sla_ws.append(["Priority", "Target (hrs)", "Avg resolution (hrs)"])
    _header_row(sla_ws, 1)
    for p, target in sla["targets_hours"].items():
        avg = sla["avg_resolution_hours_by_priority"].get(p, "—")
        sla_ws.append([p, target, avg])
    _autosize(sla_ws)

    # --- Trend sheet ---
    trend_ws = wb.create_sheet("Trend (14 days)")
    trend_ws.append(["Date", "Created", "Resolved"])
    _header_row(trend_ws, 1)
    for row in data["trend"]:
        trend_ws.append([row["date"], row["created"], row["resolved"]])
    _autosize(trend_ws)

    # --- Raw tickets sheet ---
    tickets_ws = wb.create_sheet("Tickets")
    tickets_ws.append([
        "Ticket #", "Title", "Type", "Category", "Priority", "Status",
        "Creator", "Assignee", "Created At (UTC)", "Updated At (UTC)",
        "Resolved At (UTC)", "Closed At (UTC)",
    ])
    _header_row(tickets_ws, 1)
    tickets = (
        db.query(Ticket)
        .options(joinedload(Ticket.creator), joinedload(Ticket.assignee))
        .order_by(Ticket.created_at.desc())
        .limit(MAX_TICKET_ROWS)
        .all()
    )

    def _fmt(dt):
        return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

    for t in tickets:
        tickets_ws.append([
            t.ticket_number, t.title, t.ticket_type, t.category, t.priority, t.status,
            t.creator.display_name if t.creator else "",
            t.assignee.display_name if t.assignee else "",
            _fmt(t.created_at), _fmt(t.updated_at), _fmt(t.resolved_at), _fmt(t.closed_at),
        ])
    _autosize(tickets_ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# --------------------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------------------

def _styled_table(rows):
    table = Table(rows, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_HEX}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{BORDER_HEX}")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6F8")]),
    ]))
    return table


def build_pdf(db: Session) -> io.BytesIO:
    data = service.overview(db)
    sla = data["sla"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("EpicTitle", parent=styles["Title"], textColor=colors.HexColor(f"#{BRAND_HEX}"))
    h2_style = ParagraphStyle("EpicH2", parent=styles["Heading2"], spaceBefore=14, textColor=colors.HexColor("#1F2328"))
    muted_style = ParagraphStyle("EpicMuted", parent=styles["Normal"], textColor=colors.HexColor(f"#{MUTED_HEX}"))

    elements = [
        Paragraph("EPIC Helpdesk — Admin Dashboard Report", title_style),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", muted_style),
        Spacer(1, 14),
        Paragraph("Summary", h2_style),
        _styled_table([
            ["Metric", "Value"],
            ["Total tickets", data["total_tickets"]],
            ["Open tickets", data["open_tickets"]],
            ["SLA compliance", f'{sla["compliance_pct"]}%' if sla["compliance_pct"] is not None else "—"],
            ["SLA breached (resolved late)", sla["breached_count"]],
            ["Open tickets breaching SLA", sla["open_breached_count"]],
            ["Open tickets at risk", sla["open_at_risk_count"]],
            ["Avg resolution time (hrs)", sla["avg_resolution_hours_overall"] or "—"],
        ]),
    ]

    def breakdown_block(title: str, mapping: dict):
        elements.append(Paragraph(title, h2_style))
        rows = [["Key", "Count"]] + [[k, v] for k, v in mapping.items()]
        elements.append(_styled_table(rows))

    breakdown_block("By status", data["by_status"])
    breakdown_block("By priority", data["by_priority"])
    breakdown_block("By category", data["by_category"])
    breakdown_block("By ticket type", data["by_ticket_type"])

    elements.append(Paragraph("SLA targets &amp; average resolution", h2_style))
    sla_rows = [["Priority", "Target (hrs)", "Avg resolution (hrs)"]]
    for p, target in sla["targets_hours"].items():
        avg = sla["avg_resolution_hours_by_priority"].get(p, "—")
        sla_rows.append([p, target, avg])
    elements.append(_styled_table(sla_rows))

    elements.append(Paragraph("Ticket volume — last 14 days", h2_style))
    trend_rows = [["Date", "Created", "Resolved"]] + [
        [r["date"], r["created"], r["resolved"]] for r in data["trend"]
    ]
    elements.append(_styled_table(trend_rows))

    doc.build(elements)
    buf.seek(0)
    return buf