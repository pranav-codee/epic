"""Aggregation queries for dashboards and management reporting, plus Excel/PDF export."""
import io
from datetime import datetime, timedelta
from sqlalchemy import func
from sqlalchemy.orm import Session, aliased
from ..tickets.models import Ticket, PRIORITIES
from ..tickets import workflow as wf
from ..catalogue.models import AssignmentGroup
from ..users.models import UserProfile
from .models import DailyGroupSnapshot
from ...core.sla import SLA_HOURS_BY_PRIORITY, AT_RISK_THRESHOLD, BUSINESS_HOURS_SLA_PRIORITY_LEVELS
from ...core.time import utcnow

# SPEC §4: rolling per-priority SLA adherence-% target (P1 99%, P2 98%, P3 97%, P4 95%),
# keyed by the same P1-P4 levels BUSINESS_HOURS_SLA_PRIORITY_LEVELS (core/sla.py) maps
# CRITICAL/HIGH/MEDIUM/LOW onto. No adherence-% constant existed anywhere yet (core/sla.py
# only had the response/resolution time targets, not the rolling compliance-% target on top
# of them), so it's added here rather than invented ad hoc inside the aggregation function.
SLA_ADHERENCE_TARGET_PCT = {"P1": 99, "P2": 98, "P3": 97, "P4": 95}

OPEN_STATES = ("OPEN", "ASSIGNED", "IN_PROGRESS", "PENDING_USER")
TREND_DAYS = 14


def _sla_breakdown(db: Session, now: datetime | None = None):
    """Single pass over (priority, created_at, resolved_at, closed_at, sla_due_at,
    sla_at_risk_notified_at, sla_breached_notified_at) to derive SLA compliance, breach,
    at-risk, and escalation-notification counts without pulling full Ticket rows.

    The escalation counts (`at_risk_escalated` / `breached_escalated`) reflect
    app.core.sla_scanner activity: they count tickets that have ever had that
    notification fire (the notified-at column is non-NULL), not just tickets
    currently sitting in that state. This is what makes it possible for the
    dashboard to show "escalations sent" as distinct from "currently at risk" —
    e.g. a ticket can be currently AT_RISK but not yet escalated (scanner hasn't
    ticked yet) or no longer AT_RISK but still show as having been escalated at
    some point during its lifetime.
    """
    now = now or utcnow()
    rows = db.query(
        Ticket.priority, Ticket.created_at, Ticket.resolved_at, Ticket.closed_at, Ticket.sla_due_at,
        Ticket.sla_at_risk_notified_at, Ticket.sla_breached_notified_at,
    ).all()

    met = breached = at_risk = on_track = no_sla = 0
    resolved_count = 0
    total_resolution_seconds = 0.0
    by_priority_resolution = {p: {"count": 0, "seconds": 0.0} for p in PRIORITIES}
    at_risk_escalated = 0
    breached_escalated = 0
    last_escalation_at = None

    for priority, created_at, resolved_at, closed_at, sla_due_at, at_risk_notified_at, breached_notified_at in rows:
        if at_risk_notified_at is not None:
            at_risk_escalated += 1
            if last_escalation_at is None or at_risk_notified_at > last_escalation_at:
                last_escalation_at = at_risk_notified_at
        if breached_notified_at is not None:
            breached_escalated += 1
            if last_escalation_at is None or breached_notified_at > last_escalation_at:
                last_escalation_at = breached_notified_at

        end = resolved_at or closed_at
        if end is not None:
            resolved_count += 1
            seconds = (end - created_at).total_seconds()
            total_resolution_seconds += seconds
            if priority in by_priority_resolution:
                by_priority_resolution[priority]["count"] += 1
                by_priority_resolution[priority]["seconds"] += seconds

        if sla_due_at is None:
            no_sla += 1
            continue
        if end is not None:
            if end <= sla_due_at:
                met += 1
            else:
                breached += 1
            continue
        if now > sla_due_at:
            breached += 1
            continue
        window = (sla_due_at - created_at).total_seconds()
        remaining = (sla_due_at - now).total_seconds()
        if window > 0 and (remaining / window) <= AT_RISK_THRESHOLD:
            at_risk += 1
        else:
            on_track += 1

    evaluated = met + breached
    compliance_rate = round((met / evaluated) * 100, 1) if evaluated else None
    avg_resolution_hours = (
        round((total_resolution_seconds / resolved_count) / 3600, 1) if resolved_count else None
    )
    avg_resolution_by_priority = {
        p: (round((v["seconds"] / v["count"]) / 3600, 1) if v["count"] else None)
        for p, v in by_priority_resolution.items()
    }

    return {
        "compliance_rate": compliance_rate,
        "met": met,
        "breached": breached,
        "at_risk": at_risk,
        "on_track": on_track,
        "no_sla_target": no_sla,
        "avg_resolution_hours": avg_resolution_hours,
        "avg_resolution_hours_by_priority": avg_resolution_by_priority,
        "target_hours_by_priority": SLA_HOURS_BY_PRIORITY,
        # NEW — app.core.sla_scanner escalation activity (historical, not point-in-time).
        "escalations": {
            "at_risk_notified": at_risk_escalated,
            "breached_notified": breached_escalated,
            "total_notified": at_risk_escalated + breached_escalated,
            "last_escalation_at": last_escalation_at.isoformat() if last_escalation_at else None,
        },
    }


def _trend(db: Session, days: int = TREND_DAYS, now: datetime | None = None):
    now = now or utcnow()
    since = (now - timedelta(days=days - 1)).replace(hour=0, minute=0, second=0, microsecond=0)
    rows = (
        db.query(func.date(Ticket.created_at), func.count(Ticket.id))
        .filter(Ticket.created_at >= since)
        .group_by(func.date(Ticket.created_at))
        .all()
    )
    counts = {str(d): c for d, c in rows}
    out = []
    for i in range(days):
        day = (since + timedelta(days=i)).date()
        out.append({"date": day.isoformat(), "count": counts.get(day.isoformat(), 0)})
    return out


def overview(db: Session):
    by_status = dict(db.query(Ticket.status, func.count(Ticket.id)).group_by(Ticket.status).all())
    by_ticket_type = dict(db.query(Ticket.ticket_type, func.count(Ticket.id)).group_by(Ticket.ticket_type).all())
    by_category = dict(db.query(Ticket.category, func.count(Ticket.id)).group_by(Ticket.category).all())
    by_priority = dict(db.query(Ticket.priority, func.count(Ticket.id)).group_by(Ticket.priority).all())
    total = db.query(func.count(Ticket.id)).scalar() or 0
    open_total = db.query(func.count(Ticket.id)).filter(Ticket.status.in_(OPEN_STATES)).scalar() or 0
    return {
        "total_tickets": total,
        "open_tickets": open_total,
        "by_status": by_status,
        "by_ticket_type": by_ticket_type,
        "by_category": by_category,
        "by_priority": by_priority,
        "sla": _sla_breakdown(db),
        "trend": _trend(db),
    }


def sla_adherence_by_priority(db: Session):
    """SPEC §4: rolling per-priority SLA adherence — for each priority (CRITICAL/HIGH/
    MEDIUM/LOW) and each independent clock (response/resolution), returns achieved count,
    breached count, total (evaluated) count, achieved %, and the target % (P1 99/P2 98/
    P3 97/P4 95, via SLA_ADHERENCE_TARGET_PCT above).

    Sourced from the ticket-level `response_sla_status`/`resolution_sla_status` fields
    (tickets/models.py) — the one-time MET/BREACHED verdict the business-hours SLA engine
    writes at first response / resolution (tickets/service.py's `_apply_sla_evaluation`) —
    which is a different, newer source of truth than the legacy wall-clock `sla_due_at`
    fields `_sla_breakdown()` above still uses. A ticket whose clock hasn't fired yet (still
    NULL — not yet first-responded-to / not yet resolved) contributes to neither achieved
    nor breached for that clock; only evaluated tickets count toward the total and the
    achieved %, consistent with how `_sla_breakdown()`'s own `compliance_rate` already
    excludes not-yet-evaluated tickets.
    """
    rows = db.query(Ticket.priority, Ticket.response_sla_status, Ticket.resolution_sla_status).all()

    clocks = ("response", "resolution")
    counts = {p: {c: {"achieved": 0, "breached": 0} for c in clocks} for p in PRIORITIES}

    for priority, response_status, resolution_status in rows:
        if priority not in counts:
            # Defensive: `priority` has no DB-level check constraint, so guard against a
            # stale/unexpected value rather than KeyError on it.
            continue
        if response_status == "MET":
            counts[priority]["response"]["achieved"] += 1
        elif response_status == "BREACHED":
            counts[priority]["response"]["breached"] += 1
        if resolution_status == "MET":
            counts[priority]["resolution"]["achieved"] += 1
        elif resolution_status == "BREACHED":
            counts[priority]["resolution"]["breached"] += 1

    result = {}
    for priority in PRIORITIES:
        level = BUSINESS_HOURS_SLA_PRIORITY_LEVELS.get(priority)
        target_pct = SLA_ADHERENCE_TARGET_PCT.get(level)
        by_clock = {}
        for clock in clocks:
            achieved = counts[priority][clock]["achieved"]
            breached = counts[priority][clock]["breached"]
            total = achieved + breached
            achieved_pct = round((achieved / total) * 100, 1) if total else None
            by_clock[clock] = {
                "achieved": achieved,
                "breached": breached,
                "total": total,
                "achieved_pct": achieved_pct,
                "target_pct": target_pct,
            }
        result[priority] = {"level": level, **by_clock}

    return result


def breached_tickets_detail(db: Session, ticket_type: str, sla_clock: str):
    """Production View B drill-down: every ticket of `ticket_type` currently BREACHED on
    the given SLA clock ("response" -> Ticket.response_sla_status, "resolution" ->
    Ticket.resolution_sla_status — the same one-time MET/BREACHED verdict columns
    sla_adherence_by_priority() above reads), with the fields the drill-down table needs:
    ticket number, created_at, title, priority, assignment group name, technician
    (assignee) display name, and the free-text breached_reason (tickets/models.py —
    required app-side once either SLA status is BREACHED).

    Not filtered by open/closed status: a ticket keeps its BREACHED verdict permanently
    once that clock has fired (sla_adherence_by_priority() counts it the same way), so a
    since-resolved ticket that still breached its clock stays in this list — the point of
    the drill-down is "which tickets breached", not "which tickets are still open".

    assignment_group / technician are None when the ticket has no assignment_group_id /
    assignee_id respectively, rather than being coerced to a placeholder string — this is
    a raw per-ticket detail list, not a crosstab that needs a stable "Unassigned" row.
    """
    if ticket_type not in wf.WORKFLOW_ENABLED_TICKET_TYPES:
        raise ValueError(f"Invalid ticket_type. Allowed: {sorted(wf.WORKFLOW_ENABLED_TICKET_TYPES)}")
    if sla_clock not in ("response", "resolution"):
        raise ValueError("Invalid sla_clock. Allowed: response, resolution")

    status_col = Ticket.response_sla_status if sla_clock == "response" else Ticket.resolution_sla_status

    Assignee = aliased(UserProfile)
    rows = (
        db.query(
            Ticket.ticket_number,
            Ticket.created_at,
            Ticket.title,
            Ticket.priority,
            AssignmentGroup.name,
            Assignee.display_name,
            Ticket.breached_reason,
        )
        .outerjoin(AssignmentGroup, Ticket.assignment_group_id == AssignmentGroup.id)
        .outerjoin(Assignee, Ticket.assignee_id == Assignee.id)
        .filter(Ticket.ticket_type == ticket_type, status_col == "BREACHED")
        .order_by(Ticket.created_at.asc())
        .all()
    )

    return [
        {
            "ticket_number": ticket_number,
            "created_at": created_at.isoformat() if created_at else None,
            "title": title,
            "priority": priority,
            "assignment_group": group_name,
            "technician": assignee_name,
            "breached_reason": breached_reason,
        }
        for ticket_number, created_at, title, priority, group_name, assignee_name, breached_reason in rows
    ]


def sla_compliance_view(db: Session, ticket_type: str):
    """Production View B: combines sla_adherence_by_priority()'s achieved%/target% matrix
    with the breached-tickets drill-down for both clocks, scoped to a single ticket_type
    (the view is "for each ticket type" per spec — sla_adherence_by_priority() itself isn't
    ticket_type-scoped since it's used elsewhere as a global rollup, so the per-type
    breached-ticket filtering happens here instead).
    """
    if ticket_type not in wf.WORKFLOW_ENABLED_TICKET_TYPES:
        raise ValueError(f"Invalid ticket_type. Allowed: {sorted(wf.WORKFLOW_ENABLED_TICKET_TYPES)}")

    return {
        "ticket_type": ticket_type,
        "adherence_by_priority": sla_adherence_by_priority(db),
        "breached_tickets": {
            "response": breached_tickets_detail(db, ticket_type, "response"),
            "resolution": breached_tickets_detail(db, ticket_type, "resolution"),
        },
    }


def inflow_resolved_open_by_group(db: Session, now: datetime | None = None):
    """Counts of tickets created / resolved / currently open, grouped by
    Ticket.assignment_group_id (joined to AssignmentGroup.name), for the current period.

    "Current period" = current calendar month to date (1st of month 00:00 UTC through
    `now`), matching the "for the current period" reporting cadence used elsewhere in this
    module. `created` and `resolved` are period-scoped (only tickets created / resolved
    within the window count); `open` is a point-in-time snapshot of tickets currently in
    OPEN_STATES, independent of the period, since "currently open" isn't a period-bounded
    concept. Tickets with no assignment_group_id are rolled up under "Unassigned" rather
    than dropped, so the sheet's totals still reconcile against the ticket count.
    """
    now = now or utcnow()
    period_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    group_name = func.coalesce(AssignmentGroup.name, "Unassigned")

    def _grouped(q):
        return dict(
            q.outerjoin(AssignmentGroup, Ticket.assignment_group_id == AssignmentGroup.id)
            .group_by(group_name)
            .all()
        )

    created = _grouped(
        db.query(group_name, func.count(Ticket.id)).filter(
            Ticket.created_at >= period_start, Ticket.created_at <= now
        )
    )
    resolved_at = func.coalesce(Ticket.resolved_at, Ticket.closed_at)
    resolved = _grouped(
        db.query(group_name, func.count(Ticket.id)).filter(
            resolved_at.isnot(None), resolved_at >= period_start, resolved_at <= now
        )
    )
    open_now = _grouped(
        db.query(group_name, func.count(Ticket.id)).filter(Ticket.status.in_(OPEN_STATES))
    )

    names = sorted(set(created) | set(resolved) | set(open_now))
    return {
        "period_start": period_start.isoformat(),
        "period_end": now.isoformat(),
        "by_group": [
            {
                "group": name,
                "created": created.get(name, 0),
                "resolved": resolved.get(name, 0),
                "open": open_now.get(name, 0),
            }
            for name in names
        ],
    }


_DAILY_OPS_TICKET_TYPES = ("INCIDENT", "SERVICE_REQUEST")


def take_daily_snapshot(db: Session, snapshot_date=None) -> int:
    """Writes one DailyGroupSnapshot row per Assignment Group (plus one NULL-group row for
    the "Unassigned" bucket), capturing each group's current open Incident / Service
    Request counts. Intended to be called once per day by app.core.daily_snapshot_loop, but
    also directly callable (e.g. for backfill or tests) with an explicit snapshot_date.

    This is the mechanism daily_ops_summary()'s "Yesterday's Backlog" column reads from —
    without a persisted point-in-time snapshot, "yesterday's open-ticket count" is
    unrecoverable once today's tickets have already mutated the live OPEN_STATES-filtered
    count.

    Idempotent per snapshot_date: any existing rows for that date are deleted and replaced
    with a freshly computed full set in the same transaction, so re-running for a date that
    already has rows (a loop restart mid-day, a manual backfill re-run, or two app instances
    both firing the same day) converges on one consistent end state rather than
    accumulating duplicates.
    """
    snapshot_date = snapshot_date or utcnow().date()

    open_counts = (
        db.query(Ticket.assignment_group_id, Ticket.ticket_type, func.count(Ticket.id))
        .filter(Ticket.status.in_(OPEN_STATES))
        .group_by(Ticket.assignment_group_id, Ticket.ticket_type)
        .all()
    )

    by_group: dict = {}
    for group_id, ticket_type, count in open_counts:
        bucket = by_group.setdefault(group_id, {"INCIDENT": 0, "SERVICE_REQUEST": 0})
        if ticket_type in bucket:
            bucket[ticket_type] = count

    # Every Assignment Group gets a row even if it currently has zero open tickets, so a
    # later "no snapshot row for this group" is distinguishable from "snapshot row exists
    # and says zero" — the former means the job never ran for that group, not that its
    # backlog was empty. The NULL-group ("Unassigned") bucket is always included too, even
    # if empty, for the same reason.
    for (group_id,) in db.query(AssignmentGroup.id).all():
        by_group.setdefault(group_id, {"INCIDENT": 0, "SERVICE_REQUEST": 0})
    by_group.setdefault(None, {"INCIDENT": 0, "SERVICE_REQUEST": 0})

    db.query(DailyGroupSnapshot).filter(
        DailyGroupSnapshot.snapshot_date == snapshot_date
    ).delete(synchronize_session=False)

    for group_id, counts in by_group.items():
        db.add(DailyGroupSnapshot(
            snapshot_date=snapshot_date,
            assignment_group_id=group_id,
            open_incidents_count=counts["INCIDENT"],
            open_srs_count=counts["SERVICE_REQUEST"],
        ))
    db.commit()
    return len(by_group)


def daily_ops_summary(db: Session, now: datetime | None = None):
    """Production View A: per Assignment Group, a day-over-day delta — Inflow (Incidents/
    SRs created today), Closures (Incidents/SRs resolved or fulfilled today), Yesterday's
    Backlog (from the DailyGroupSnapshot taken for yesterday's date), and Today's Backlog
    (current open count). Each of the four metrics is split by ticket_type (Incidents vs
    Service Requests), matching production's layout.

    "Today" = current calendar day to date (00:00 UTC through `now`); "yesterday" = the
    calendar day immediately before that. Inflow/Closures are period-scoped (only today's
    creations/resolutions count); Today's Backlog is a point-in-time snapshot of currently-
    open tickets, independent of the period — same conventions
    inflow_resolved_open_by_group() above already uses for its month-scoped equivalents.

    Yesterday's Backlog comes from the persisted DailyGroupSnapshot table rather than a live
    query, since "yesterday's open count" can no longer be reconstructed live once today's
    tickets have already changed the open-ticket set (see take_daily_snapshot() above). If
    the snapshot job hasn't run yet for yesterday's date (e.g. first day after deploy), a
    group simply shows 0 for that column rather than raising — the same "missing data reads
    as zero" behavior the rest of this module uses for groups with no matching rows.

    Tickets with no assignment_group_id are rolled up under "Unassigned", consistent with
    inflow_resolved_open_by_group() and group_by_status_crosstab() above.
    """
    now = now or utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_date = (today_start - timedelta(days=1)).date()

    group_name = func.coalesce(AssignmentGroup.name, "Unassigned")

    def _by_group_and_type(q):
        rows = (
            q.outerjoin(AssignmentGroup, Ticket.assignment_group_id == AssignmentGroup.id)
            .group_by(group_name, Ticket.ticket_type)
            .all()
        )
        out: dict = {}
        for name, ticket_type, count in rows:
            bucket = out.setdefault(name, {"INCIDENT": 0, "SERVICE_REQUEST": 0})
            if ticket_type in bucket:
                bucket[ticket_type] = count
        return out

    inflow = _by_group_and_type(
        db.query(group_name, Ticket.ticket_type, func.count(Ticket.id)).filter(
            Ticket.created_at >= today_start,
            Ticket.created_at <= now,
            Ticket.ticket_type.in_(_DAILY_OPS_TICKET_TYPES),
        )
    )

    resolved_at = func.coalesce(Ticket.resolved_at, Ticket.closed_at)
    closures = _by_group_and_type(
        db.query(group_name, Ticket.ticket_type, func.count(Ticket.id)).filter(
            resolved_at.isnot(None),
            resolved_at >= today_start,
            resolved_at <= now,
            Ticket.ticket_type.in_(_DAILY_OPS_TICKET_TYPES),
        )
    )

    today_backlog = _by_group_and_type(
        db.query(group_name, Ticket.ticket_type, func.count(Ticket.id)).filter(
            Ticket.status.in_(OPEN_STATES),
            Ticket.ticket_type.in_(_DAILY_OPS_TICKET_TYPES),
        )
    )

    snapshot_rows = (
        db.query(group_name, DailyGroupSnapshot.open_incidents_count, DailyGroupSnapshot.open_srs_count)
        .outerjoin(AssignmentGroup, DailyGroupSnapshot.assignment_group_id == AssignmentGroup.id)
        .filter(DailyGroupSnapshot.snapshot_date == yesterday_date)
        .all()
    )
    yesterday_backlog = {
        name: {"INCIDENT": inc_count, "SERVICE_REQUEST": sr_count}
        for name, inc_count, sr_count in snapshot_rows
    }

    def _get(mapping, name):
        return mapping.get(name, {"INCIDENT": 0, "SERVICE_REQUEST": 0})

    names = sorted(set(inflow) | set(closures) | set(today_backlog) | set(yesterday_backlog))
    by_group = []
    for name in names:
        inf = _get(inflow, name)
        clo = _get(closures, name)
        yb = _get(yesterday_backlog, name)
        tb = _get(today_backlog, name)
        by_group.append({
            "group": name,
            "inflow": {"incidents": inf["INCIDENT"], "srs": inf["SERVICE_REQUEST"]},
            "closures": {"incidents": clo["INCIDENT"], "srs": clo["SERVICE_REQUEST"]},
            "yesterday_backlog": {"incidents": yb["INCIDENT"], "srs": yb["SERVICE_REQUEST"]},
            "today_backlog": {"incidents": tb["INCIDENT"], "srs": tb["SERVICE_REQUEST"]},
        })

    return {
        "today_date": today_start.date().isoformat(),
        "yesterday_date": yesterday_date.isoformat(),
        "by_group": by_group,
    }


def group_by_status_crosstab(db: Session, ticket_type: str):
    """Production View D: currently-open tickets cross-tabbed by Assignment Group x
    workflow_status, computed separately per ticket_type (INCIDENT vs SERVICE_REQUEST)
    since the two have distinct workflow_status vocabularies (tickets/workflow.py's
    INCIDENT_WORKFLOW_STATUSES vs SERVICE_REQUEST_WORKFLOW_STATUSES).

    "Open" means Ticket.status in OPEN_STATES (same definition used everywhere else in this
    module) — this also happens to exclude the terminal workflow_status values (RESOLVED/
    FULFILLED) without needing a separate filter, since a ticket reaching either of those
    always has its generic `status` moved to RESOLVED too (tickets/service.py).

    Every workflow_status column for the ticket type is always present in each row's
    `counts_by_workflow_status` (zero-filled), even if no group currently has a ticket in
    that state, so the frontend can render a stable set of columns rather than a ragged
    one that changes shape depending on what data happens to exist right now.
    """
    if ticket_type not in wf.WORKFLOW_ENABLED_TICKET_TYPES:
        raise ValueError(f"Invalid ticket_type. Allowed: {sorted(wf.WORKFLOW_ENABLED_TICKET_TYPES)}")

    statuses = wf.workflow_statuses_for(ticket_type)
    group_name = func.coalesce(AssignmentGroup.name, "Unassigned")

    rows = (
        db.query(group_name, Ticket.workflow_status, func.count(Ticket.id))
        .outerjoin(AssignmentGroup, Ticket.assignment_group_id == AssignmentGroup.id)
        .filter(Ticket.ticket_type == ticket_type, Ticket.status.in_(OPEN_STATES))
        .group_by(group_name, Ticket.workflow_status)
        .all()
    )

    by_group: dict[str, dict[str, int]] = {}
    for name, workflow_status, count in rows:
        bucket = by_group.setdefault(name, {s: 0 for s in statuses})
        if workflow_status in bucket:
            bucket[workflow_status] = count
        # A row with a workflow_status outside this ticket type's set (e.g. NULL, or stale
        # data predating the workflow_status column) doesn't fit any column — it still
        # created the group's entry above via setdefault, just with no count attributed.

    return [
        {"assignment_group_name": name, "counts_by_workflow_status": by_group[name]}
        for name in sorted(by_group)
    ]


# Production View C: fixed ageing-bucket boundaries, in days-old, applied to currently-open
# tickets. Buckets are cumulative-exclusive (each ticket lands in exactly one) with the
# upper bucket open-ended: <=1, (1,3], (3,7], (7,15], (15,30], >30. Exposed as a fixed,
# ordered column set (like group_by_status_crosstab's workflow_status columns) so the
# frontend always renders the same six columns regardless of which buckets currently have
# tickets in them.
AGEING_BUCKETS = ["<=1", ">1", ">3", ">7", ">15", ">30"]


def _ageing_bucket_for(age_days: float) -> str:
    if age_days <= 1:
        return "<=1"
    if age_days <= 3:
        return ">1"
    if age_days <= 7:
        return ">3"
    if age_days <= 15:
        return ">7"
    if age_days <= 30:
        return ">15"
    return ">30"


def ageing_buckets(db: Session, ticket_type: str, channel_filter: str = "all", now: datetime | None = None):
    """Production View C: currently-open tickets bucketed by (now - created_at) age, in
    fixed <=1/>1/>3/>7/>15/>30-day buckets (AGEING_BUCKETS above), cross-tabbed by
    Assignment Group. Computed separately per ticket_type (INCIDENT vs SERVICE_REQUEST,
    same restriction as group_by_status_crosstab's ticket types) and, within ticket_type,
    separately sliced by channel:

    - channel_filter="monitoring": only channel == "MONITORING_TOOL" tickets (monitoring-tool
      alerts, e.g. PRTG).
    - channel_filter="human": only channel != "MONITORING_TOOL" tickets (end-user-raised
      tickets — email/phone/self-service).
    - channel_filter="all": no channel filter (used for Service Requests, which don't
      originate from the monitoring-tool channel).

    "Open" uses the same OPEN_STATES definition as the rest of this module. Every bucket
    column is always present in each row's `counts_by_bucket` (zero-filled), and every
    assignment group with at least one open ticket in scope gets a row — tickets with no
    assignment_group_id roll up under "Unassigned" rather than being dropped, matching
    inflow_resolved_open_by_group()'s convention.
    """
    if ticket_type not in wf.WORKFLOW_ENABLED_TICKET_TYPES:
        raise ValueError(f"Invalid ticket_type. Allowed: {sorted(wf.WORKFLOW_ENABLED_TICKET_TYPES)}")
    if channel_filter not in ("monitoring", "human", "all"):
        raise ValueError("Invalid channel. Allowed: monitoring, human, all")

    now = now or utcnow()
    group_name = func.coalesce(AssignmentGroup.name, "Unassigned")

    q = (
        db.query(group_name, Ticket.created_at)
        .outerjoin(AssignmentGroup, Ticket.assignment_group_id == AssignmentGroup.id)
        .filter(Ticket.ticket_type == ticket_type, Ticket.status.in_(OPEN_STATES))
    )
    if channel_filter == "monitoring":
        q = q.filter(Ticket.channel == "MONITORING_TOOL")
    elif channel_filter == "human":
        q = q.filter(Ticket.channel != "MONITORING_TOOL")

    by_group: dict[str, dict[str, int]] = {}
    for name, created_at in q.all():
        age_days = (now - created_at).total_seconds() / 86400
        bucket = _ageing_bucket_for(age_days)
        row = by_group.setdefault(name, {b: 0 for b in AGEING_BUCKETS})
        row[bucket] += 1

    return [
        {"assignment_group_name": name, "counts_by_bucket": by_group[name]}
        for name in sorted(by_group)
    ]


def engineer_workload(db: Session, engineer_id: str):
    rows = (db.query(Ticket.status, func.count(Ticket.id))
              .filter(Ticket.assignee_id == engineer_id)
              .group_by(Ticket.status).all())
    return {
        "engineer_id": engineer_id,
        "by_status": dict(rows),
        "active": sum(c for s, c in rows if s in ("ASSIGNED", "IN_PROGRESS", "PENDING_USER")),
    }


def report_group(db: Session, group_by: str, from_dt=None, to_dt=None):
    col_map = {"status": Ticket.status, "ticket_type": Ticket.ticket_type, "category": Ticket.category, "priority": Ticket.priority}
    col = col_map.get(group_by)
    if col is None:
        raise ValueError(f"group_by must be one of {list(col_map)}")
    q = db.query(col, func.count(Ticket.id)).group_by(col)
    if from_dt:
        q = q.filter(Ticket.created_at >= from_dt)
    if to_dt:
        q = q.filter(Ticket.created_at <= to_dt)
    return [{"key": k, "count": v} for k, v in q.all()]


# ---------- Export ----------

def _label(key: str) -> str:
    return str(key).replace("_", " ").title()


def build_excel_export(db: Session) -> bytes:
    """Builds an .xlsx workbook (Summary / By Status / By Priority / By Category /
    By Type / SLA / Trend / Inflow-Resolved-Open by Group / SLA Met-Breached by Priority
    sheets) from the same overview() data the dashboard shows, plus the group- and
    priority-level aggregations from inflow_resolved_open_by_group() and
    sla_adherence_by_priority()."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    data = overview(db)
    wb = Workbook()

    header_fill = PatternFill(start_color="0067B8", end_color="0067B8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True, color="0067B8")

    def style_header_row(ws, row_idx=1, ncols=2):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left")

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "EPIC Admin Dashboard — Summary"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated {utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    ws["A2"].font = Font(italic=True, color="57606A")

    ws.append([])
    ws.append(["Metric", "Value"])
    style_header_row(ws, row_idx=4)
    sla = data["sla"]
    esc = sla["escalations"]
    summary_rows = [
        ("Total tickets", data["total_tickets"]),
        ("Open tickets", data["open_tickets"]),
        ("SLA compliance rate (%)", sla["compliance_rate"]),
        ("SLA breached", sla["breached"]),
        ("SLA at risk", sla["at_risk"]),
        ("SLA on track", sla["on_track"]),
        ("Avg. resolution time (hours)", sla["avg_resolution_hours"]),
        ("SLA escalations sent (at-risk)", esc["at_risk_notified"]),
        ("SLA escalations sent (breached)", esc["breached_notified"]),
    ]
    for label, value in summary_rows:
        ws.append([label, value])
    autosize(ws, [32, 18])

    # --- Generic breakdown sheet helper ---
    def add_breakdown_sheet(name, mapping):
        s = wb.create_sheet(name)
        s.append([name.replace("By ", ""), "Count"])
        style_header_row(s)
        for k, v in mapping.items():
            s.append([_label(k), v])
        autosize(s, [26, 12])

    add_breakdown_sheet("By Status", data["by_status"])
    add_breakdown_sheet("By Priority", data["by_priority"])
    add_breakdown_sheet("By Category", data["by_category"])
    add_breakdown_sheet("By Type", data["by_ticket_type"])

    # --- SLA sheet ---
    sla_ws = wb.create_sheet("SLA")
    sla_ws.append(["SLA metric", "Value"])
    style_header_row(sla_ws)
    for label, value in [
        ("Compliance rate (%)", sla["compliance_rate"]),
        ("Met", sla["met"]),
        ("Breached", sla["breached"]),
        ("At risk", sla["at_risk"]),
        ("On track", sla["on_track"]),
        ("No SLA target", sla["no_sla_target"]),
        ("Avg. resolution (hours)", sla["avg_resolution_hours"]),
    ]:
        sla_ws.append([label, value])
    sla_ws.append([])
    esc = sla["escalations"]
    sla_ws.append(["SLA escalations sent (app.core.sla_scanner)", ""])
    sla_ws.append(["At-risk notifications sent", esc["at_risk_notified"]])
    sla_ws.append(["Breached notifications sent", esc["breached_notified"]])
    sla_ws.append(["Total notifications sent", esc["total_notified"]])
    sla_ws.append(["Most recent escalation (UTC)", esc["last_escalation_at"] or "—"])
    sla_ws.append([])
    sla_ws.append(["Target (hours) by priority", ""])
    for p, h in sla["target_hours_by_priority"].items():
        sla_ws.append([_label(p), h])
    sla_ws.append([])
    sla_ws.append(["Avg. resolution (hours) by priority", ""])
    for p, h in sla["avg_resolution_hours_by_priority"].items():
        sla_ws.append([_label(p), h])
    autosize(sla_ws, [32, 14])

    # --- Trend sheet ---
    trend_ws = wb.create_sheet("Ticket Trend")
    trend_ws.append(["Date", "Tickets created"])
    style_header_row(trend_ws)
    for point in data["trend"]:
        trend_ws.append([point["date"], point["count"]])
    autosize(trend_ws, [16, 16])

    # --- Inflow-Resolved-Open by Group sheet ---
    group_data = inflow_resolved_open_by_group(db)
    group_ws = wb.create_sheet("Inflow-Resolved-Open by Group")
    group_ws["A1"] = "Inflow / Resolved / Open by Assignment Group"
    group_ws["A1"].font = title_font
    group_ws["A2"] = (
        f"Period: {group_data['period_start']} to {group_data['period_end']} (UTC)"
    )
    group_ws["A2"].font = Font(italic=True, color="57606A")
    group_ws.append([])
    group_ws.append(["Assignment group", "Created", "Resolved", "Currently open"])
    style_header_row(group_ws, row_idx=4, ncols=4)
    for row in group_data["by_group"]:
        group_ws.append([row["group"], row["created"], row["resolved"], row["open"]])
    autosize(group_ws, [30, 12, 12, 16])

    # --- SLA Met-Breached by Priority sheet ---
    sla_priority_data = sla_adherence_by_priority(db)
    sla_priority_ws = wb.create_sheet("SLA Met-Breached by Priority")
    sla_priority_ws.append(
        ["Priority", "Clock", "Met", "Breached", "Total evaluated", "Achieved %", "Target %"]
    )
    style_header_row(sla_priority_ws, ncols=7)
    for priority in PRIORITIES:
        info = sla_priority_data.get(priority)
        if not info:
            continue
        for clock in ("response", "resolution"):
            c = info[clock]
            sla_priority_ws.append([
                _label(priority),
                _label(clock),
                c["achieved"],
                c["breached"],
                c["total"],
                c["achieved_pct"] if c["achieved_pct"] is not None else "—",
                c["target_pct"] if c["target_pct"] is not None else "—",
            ])
    autosize(sla_priority_ws, [14, 14, 10, 10, 16, 14, 12])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf_export(db: Session) -> bytes:
    """Builds a print-friendly PDF summary of the dashboard (KPIs, SLA, breakdown tables)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    data = overview(db)
    sla = data["sla"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()
    brand = colors.HexColor("#0067B8")
    muted = colors.HexColor("#57606A")

    title_style = ParagraphStyle("TitleBrand", parent=styles["Title"], textColor=brand, fontSize=20)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], textColor=brand, spaceBefore=14, spaceAfter=6)
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], textColor=muted)

    story = [
        Paragraph("EPIC Admin Dashboard Report", title_style),
        Paragraph(f"Generated {utcnow().strftime('%Y-%m-%d %H:%M')} UTC", meta_style),
        Spacer(1, 12),
    ]

    def make_table(rows, col_widths=(3 * inch, 2 * inch)):
        t = Table(rows, colWidths=list(col_widths))
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), brand),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F6F8")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E1E4E8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    story.append(Paragraph("Key metrics", section_style))
    story.append(make_table([
        ["Metric", "Value"],
        ["Total tickets", str(data["total_tickets"])],
        ["Open tickets", str(data["open_tickets"])],
        ["SLA compliance rate", f"{sla['compliance_rate']}%" if sla["compliance_rate"] is not None else "—"],
        ["Avg. resolution time", f"{sla['avg_resolution_hours']} hrs" if sla["avg_resolution_hours"] is not None else "—"],
    ]))

    story.append(Paragraph("SLA status", section_style))
    story.append(make_table([
        ["SLA status", "Tickets"],
        ["Met", str(sla["met"])],
        ["Breached", str(sla["breached"])],
        ["At risk", str(sla["at_risk"])],
        ["On track", str(sla["on_track"])],
        ["No SLA target", str(sla["no_sla_target"])],
    ]))

    esc = sla["escalations"]
    story.append(Paragraph("SLA escalations sent", section_style))
    story.append(make_table([
        ["Escalation type", "Notifications sent"],
        ["At-risk", str(esc["at_risk_notified"])],
        ["Breached", str(esc["breached_notified"])],
        ["Total", str(esc["total_notified"])],
        ["Most recent (UTC)", esc["last_escalation_at"] or "—"],
    ]))

    for title, mapping in [
        ("Tickets by status", data["by_status"]),
        ("Tickets by priority", data["by_priority"]),
        ("Tickets by category", data["by_category"]),
        ("Tickets by type", data["by_ticket_type"]),
    ]:
        story.append(Paragraph(title, section_style))
        rows = [["Value", "Count"]] + [[_label(k), str(v)] for k, v in mapping.items()]
        story.append(make_table(rows))

    doc.build(story)
    return buf.getvalue()